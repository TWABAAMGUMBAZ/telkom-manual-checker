from __future__ import annotations

import json
import os
import csv
import secrets
import threading
import time
import uuid
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, redirect, render_template_string, request, send_file, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

from lookup_providers import LookupProvider, build_lookup_provider
from telkom_batch_check import (
    LookupResult,
    classify_number,
    is_supported_number,
    load_cache,
    normalize_number,
    provider_is_telkom,
    save_cache,
)


DATA_DIR = Path(os.environ.get("DATA_DIR", "/tmp/telkom-render-checker"))
UPLOAD_DIR = DATA_DIR / "uploads"
REPORT_DIR = DATA_DIR / "reports"
CACHE_PATH = DATA_DIR / "telkom_lookup_cache.json"
STATE_PATH = DATA_DIR / "state.json"
JOBS_PATH = DATA_DIR / "api_jobs.json"
APP_PASSWORD = os.environ.get("APP_PASSWORD", "").strip()
API_KEYS = {key.strip() for key in os.environ.get("API_KEYS", APP_PASSWORD).replace(",", "\n").splitlines() if key.strip()}
FORM_URL = "https://www.porting.co.za/PublicWebsiteApp/#/number-inquiry?sid=smppipd4x1"

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(16))


class CloudState:
    def __init__(self) -> None:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        REPORT_DIR.mkdir(parents=True, exist_ok=True)
        self.lookup_provider: LookupProvider = build_lookup_provider()
        self.cache = load_cache(CACHE_PATH)
        self.pending_captcha: dict[str, dict[str, str]] = {}
        self.files: list[dict[str, str]] = []
        self.rows: list[dict[str, Any]] = []
        self.job_lock = threading.Lock()
        self.api_job_lock = threading.Lock()
        self.api_jobs: dict[str, dict[str, Any]] = {}
        self.auto_job: dict[str, Any] = {
            "running": False,
            "stop_requested": False,
            "checked_now": 0,
            "last_number": "",
            "last_message": "Idle",
            "blocked": False,
            "started_at": "",
            "finished_at": "",
        }
        self.load_api_jobs()
        self.load_state()

    def load_state(self) -> None:
        if STATE_PATH.exists():
            try:
                raw = json.loads(STATE_PATH.read_text(encoding="utf-8"))
                self.files = raw.get("files", [])
            except json.JSONDecodeError:
                self.files = []
                STATE_PATH.unlink(missing_ok=True)
        self.rows = self.load_rows()

    def save_state(self) -> None:
        STATE_PATH.write_text(json.dumps({"files": self.files}, indent=2), encoding="utf-8")

    def load_api_jobs(self) -> None:
        if JOBS_PATH.exists():
            try:
                self.api_jobs = json.loads(JOBS_PATH.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                self.api_jobs = {}
                JOBS_PATH.unlink(missing_ok=True)

    def save_api_jobs(self) -> None:
        JOBS_PATH.write_text(json.dumps(self.api_jobs, indent=2), encoding="utf-8")

    def reset(self) -> None:
        self.files = []
        self.rows = []
        self.cache = {}
        self.pending_captcha = {}
        for pattern in ("*.xlsx", "*.csv"):
            for path in UPLOAD_DIR.glob(pattern):
                path.unlink(missing_ok=True)
        for path in REPORT_DIR.glob("*.xlsx"):
            path.unlink(missing_ok=True)
        STATE_PATH.unlink(missing_ok=True)
        CACHE_PATH.unlink(missing_ok=True)

    def add_upload(self, file_storage) -> None:
        filename = secure_filename(file_storage.filename or "")
        if Path(filename).suffix.lower() not in {".xlsx", ".csv"}:
            raise ValueError("Please upload .xlsx or .csv files only.")
        target = UPLOAD_DIR / filename
        file_storage.save(target)
        label = target.stem.replace("_", " ").replace("-", " ").title()
        self.files = [item for item in self.files if item["path"] != str(target)]
        self.files.append({"label": label, "path": str(target)})
        self.save_state()
        self.rows = self.load_rows()

    @staticmethod
    def porting_lookup_url(clean_number: str) -> str:
        return f"{FORM_URL}&msisdn={clean_number}" if clean_number else FORM_URL

    def load_rows(self) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for item in self.files:
            path = Path(item["path"])
            if not path.exists():
                continue
            for row in self.iter_source_rows(item):
                rows.extend(self.extract_phone_rows(item, row))
        return rows

    @staticmethod
    def find_number_columns(headers: list[Any]) -> list[int]:
        matches: list[int] = []
        for idx, header in enumerate(headers, start=1):
            text = str(header or "").strip().lower()
            if text == "contact number" or ("contact" in text and "number" in text):
                matches.append(idx)
                continue
            if text in {"phone", "phone number", "telephone", "number", "msisdn", "mobile", "cell", "cellphone"}:
                matches.append(idx)
        return matches or [2]

    @staticmethod
    def find_company_column(headers: list[Any]) -> int:
        for idx, header in enumerate(headers, start=1):
            text = str(header or "").strip().lower()
            if text in {"company", "company name", "business", "business name", "name"}:
                return idx
        return 1

    @staticmethod
    def find_address_column(headers: list[Any]) -> int | None:
        address_terms = {"address", "physical address", "street address", "street", "suburb", "area", "city", "location"}
        for idx, header in enumerate(headers, start=1):
            text = str(header or "").strip().lower()
            if text in address_terms or "address" in text or text in {"suburb", "area"}:
                return idx
        return None

    def iter_source_rows(self, item: dict[str, str]) -> list[dict[str, Any]]:
        path = Path(item["path"])
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self.iter_csv_rows(path)
        return self.iter_xlsx_rows(path)

    def iter_xlsx_rows(self, path: Path) -> list[dict[str, Any]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            row_iter = ws.iter_rows(values_only=True)
            headers = list(next(row_iter, []) or [])
            number_cols = self.find_number_columns(headers)
            company_col = self.find_company_column(headers)
            address_col = self.find_address_column(headers)
            rows = []
            for row_num, values in enumerate(row_iter, start=2):
                rows.append(
                    {
                        "row": row_num,
                        "headers": headers,
                        "values": list(values or []),
                        "number_cols": number_cols,
                        "company_col": company_col,
                        "address_col": address_col,
                    }
                )
            return rows
        finally:
            wb.close()

    def iter_csv_rows(self, path: Path) -> list[dict[str, Any]]:
        text = ""
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                text = path.read_text(encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        reader = csv.reader(text.splitlines())
        rows = list(reader)
        if not rows:
            return []
        headers = rows[0]
        number_cols = self.find_number_columns(headers)
        company_col = self.find_company_column(headers)
        address_col = self.find_address_column(headers)
        return [
            {
                "row": idx,
                "headers": headers,
                "values": values,
                "number_cols": number_cols,
                "company_col": company_col,
                "address_col": address_col,
            }
            for idx, values in enumerate(rows[1:], start=2)
        ]

    def extract_phone_rows(self, item: dict[str, str], source_row: dict[str, Any]) -> list[dict[str, Any]]:
        values = source_row["values"]
        headers = source_row["headers"]
        company_col = source_row["company_col"]
        address_col = source_row.get("address_col")
        company = values[company_col - 1] if len(values) >= company_col else ""
        address = values[address_col - 1] if address_col and len(values) >= address_col else ""
        extracted: list[dict[str, Any]] = []
        seen_numbers: set[str] = set()
        for number_col in source_row["number_cols"]:
            original = values[number_col - 1] if len(values) >= number_col else ""
            clean = normalize_number(original)
            if not is_supported_number(clean) or clean in seen_numbers:
                continue
            seen_numbers.add(clean)
            header = headers[number_col - 1] if len(headers) >= number_col else "Number"
            extracted.append(
                {
                    "area": item["label"],
                    "source": item["path"],
                    "row": source_row["row"],
                    "company": company or "",
                    "address": address or "",
                    "source_column": str(header or "Number"),
                    "original_number": original or "",
                    "clean_number": clean,
                    "number_type": classify_number(clean),
                    "visit_area": self.derive_visit_area(address, item["label"], classify_number(clean)),
                }
            )
        return extracted

    @staticmethod
    def derive_visit_area(address: Any, source_label: str, number_type: str) -> str:
        text = str(address or "").strip()
        if text:
            parts = [part.strip() for part in text.replace("\n", ",").split(",") if part.strip()]
            if parts:
                return parts[-1][:80]
        if "Tshwane" in number_type:
            return "Tshwane / Pretoria"
        if "Johannesburg" in number_type:
            return "Johannesburg"
        return source_label

    def summary(self) -> dict[str, Any]:
        checked_numbers = {number for number, result in self.cache.items() if result.lookup_status == "Found"}
        checked_rows = [row for row in self.rows if row["clean_number"] in checked_numbers]
        telkom_rows = [row for row in checked_rows if self.cache[row["clean_number"]].telkom == "Yes"]
        non_telkom_rows = [row for row in checked_rows if self.cache[row["clean_number"]].telkom == "No"]
        unique_numbers = {row["clean_number"] for row in self.rows if row["clean_number"]}
        plans = [self.lead_plan(row, self.cache.get(row["clean_number"])) for row in self.rows]
        area_counts: dict[str, int] = {}
        for row in self.rows:
            area = row.get("visit_area") or row.get("area") or "Unassigned"
            area_counts[area] = area_counts.get(area, 0) + 1
        return {
            "files": len(self.files),
            "lookup_provider": self.lookup_provider.name,
            "total_rows": len(self.rows),
            "unique_numbers": len(unique_numbers),
            "checked_rows": len(checked_rows),
            "remaining_rows": len(self.rows) - len(checked_rows),
            "unique_checked_numbers": len(checked_numbers),
            "telkom_rows": len(telkom_rows),
            "non_telkom_rows": len(non_telkom_rows),
            "joburg_rows": sum(1 for row in self.rows if row.get("number_type") == "Johannesburg 011"),
            "tshwane_rows": sum(1 for row in self.rows if row.get("number_type") == "Tshwane 012"),
            "mobile_rows": sum(1 for row in self.rows if row.get("number_type") == "Mobile"),
            "hot_leads": sum(1 for plan in plans if plan["priority"] == "Hot"),
            "warm_leads": sum(1 for plan in plans if plan["priority"] == "Warm"),
            "cold_leads": sum(1 for plan in plans if plan["priority"] == "Cold"),
            "top_visit_areas": [
                {"area": area, "count": count}
                for area, count in sorted(area_counts.items(), key=lambda item: item[1], reverse=True)[:5]
            ],
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "auto_job": dict(self.auto_job),
        }

    def next_row(self) -> dict[str, Any] | None:
        for row in self.rows:
            clean = row["clean_number"]
            if clean and clean not in self.cache:
                return row
        return None

    def checked_rows(self, limit: int = 30) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        for row in self.rows:
            result = self.cache.get(row["clean_number"])
            if not result:
                continue
            items.append(
                {
                    **row,
                    "status": result.lookup_status,
                    "provider": result.current_provider,
                    "telkom": result.telkom,
                    "raw": result.raw_result,
                    "number_type": row.get("number_type", ""),
                    **self.lead_plan(row, result),
                }
            )
        return items[-limit:]

    def lead_plan(self, row: dict[str, Any], result: LookupResult | None) -> dict[str, Any]:
        score = 0
        reasons: list[str] = []
        provider = result.current_provider if result else ""
        telkom = result.telkom if result else "Unknown"
        number_type = row.get("number_type", "")

        if telkom == "Yes":
            score += 45
            reasons.append("Confirmed Telkom-service lead")
        elif telkom == "No":
            score += 35
            reasons.append(f"Competitor/provider opportunity: {provider}")
        elif result and result.lookup_status != "Found":
            score += 10
            reasons.append("Needs verification before field visit")
        else:
            score += 5
            reasons.append("Pending provider check")

        if number_type in {"Johannesburg 011", "Tshwane 012"}:
            score += 20
            reasons.append("Fixed-line business area")
        elif number_type == "Mobile":
            score += 8
            reasons.append("Mobile contact")

        if row.get("address"):
            score += 15
            reasons.append("Physical address available")
        if row.get("company"):
            score += 10
            reasons.append("Company name available")

        priority = "Cold"
        if score >= 70:
            priority = "Hot"
        elif score >= 45:
            priority = "Warm"

        if telkom == "Yes":
            action = "Visit first: confirm service, upsell, retain, or expand account."
        elif telkom == "No":
            action = "Visit for competitor conversion opportunity."
        elif result and result.lookup_status != "Found":
            action = "Verify provider manually before assigning a field visit."
        else:
            action = "Complete lookup before route planning."

        return {
            "lead_score": score,
            "priority": priority,
            "visit_area": row.get("visit_area") or row.get("area") or "Unassigned",
            "assigned_rep": "",
            "visit_status": "Not visited",
            "visit_notes": "",
            "next_action": action,
            "score_reason": "; ".join(reasons),
        }

    def save_result(self, result: LookupResult) -> None:
        if result.lookup_status == "Found":
            self.cache[result.clean_number] = result
            save_cache(CACHE_PATH, self.cache)

    def start_auto_check(self, delay_seconds: float = 1.5) -> dict[str, Any]:
        with self.job_lock:
            if self.auto_job.get("running"):
                return dict(self.auto_job)
            self.auto_job.update(
                {
                    "running": True,
                    "stop_requested": False,
                    "checked_now": 0,
                    "last_number": "",
                    "last_message": "Starting automatic checks.",
                    "blocked": False,
                    "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "finished_at": "",
                }
            )
        thread = threading.Thread(target=self.auto_check_worker, args=(delay_seconds,), daemon=True)
        thread.start()
        return dict(self.auto_job)

    def stop_auto_check(self) -> dict[str, Any]:
        with self.job_lock:
            self.auto_job["stop_requested"] = True
            self.auto_job["last_message"] = "Stopping after the current lookup."
            return dict(self.auto_job)

    def auto_check_worker(self, delay_seconds: float) -> None:
        try:
            while True:
                with self.job_lock:
                    if self.auto_job.get("stop_requested"):
                        self.auto_job["last_message"] = "Stopped by user."
                        self.pending_captcha = {}
                        break
                row = self.next_row()
                if not row:
                    with self.job_lock:
                        self.auto_job["last_message"] = "All queued numbers with automatic results are complete."
                    break
                clean_number = row["clean_number"]
                result_data = self.check_number(clean_number)
                with self.job_lock:
                    self.auto_job["last_number"] = clean_number
                    if result_data.get("kind") == "found":
                        self.auto_job["checked_now"] += 1
                        result = result_data.get("result", {})
                        self.auto_job["last_message"] = (
                            f"{clean_number}: {result.get('current_provider') or 'Unknown'} "
                            f"({result.get('telkom') or 'Unknown'})"
                        )
                    elif result_data.get("kind") == "captcha":
                        self.auto_job["blocked"] = True
                        self.auto_job["last_message"] = (
                            f"Paused at {clean_number}: type the captcha shown, then resume auto-checking."
                        )
                        break
                    else:
                        self.auto_job["blocked"] = True
                        self.auto_job["last_message"] = (
                            f"Paused at {clean_number}: manual form verification is needed."
                        )
                        break
                if delay_seconds > 0:
                    time.sleep(delay_seconds)
        finally:
            with self.job_lock:
                self.auto_job["running"] = False
                self.auto_job["stop_requested"] = False
                self.auto_job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def check_number(self, clean_number: str) -> dict[str, Any]:
        if clean_number in self.cache:
            return {"kind": "found", "result": asdict(self.cache[clean_number])}
        result = self.lookup_provider.lookup(clean_number)
        if result.lookup_status == "Found":
            self.save_result(result)
            return {"kind": "found", "result": asdict(result)}
        if result.lookup_status == "Captcha required":
            return self.new_captcha(clean_number)
        message = "Endpoint lookup did not return a confirmed provider. Use the public form fallback, then save the visible provider here."
        return {
            "kind": "manual",
            "number": clean_number,
            "formUrl": self.porting_lookup_url(clean_number),
            "result": asdict(result),
            "message": message,
        }

    def new_captcha(self, clean_number: str) -> dict[str, Any]:
        if not self.lookup_provider.supports_captcha:
            return {
                "kind": "manual",
                "number": clean_number,
                "formUrl": self.porting_lookup_url(clean_number),
                "result": asdict(
                    LookupResult(
                        clean_number,
                        "Needs review",
                        "",
                        "Unknown",
                        "The configured lookup provider does not support captcha submission.",
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    )
                ),
                "message": "Manual verification is required for this number.",
            }
        challenge = self.lookup_provider.new_captcha(clean_number)
        image_data = challenge.get("imageData", "")
        self.pending_captcha[clean_number] = {
            "captchaEncrypt": challenge.get("captchaEncrypt", ""),
            "imageData": image_data,
        }
        return {
            "kind": "captcha",
            "number": clean_number,
            "imageData": image_data,
            "message": "The Porting site requested captcha verification. Type the captcha shown, then continue automatic checking.",
        }

    def pending_captcha_payload(self) -> dict[str, str] | None:
        for number, data in self.pending_captcha.items():
            return {
                "number": number,
                "imageData": data.get("imageData", ""),
                "message": "The Porting site requested captcha verification. Type the captcha shown, then continue automatic checking.",
            }
        return None

    def submit_captcha(self, clean_number: str, code: str) -> dict[str, Any]:
        pending = self.pending_captcha.get(clean_number)
        if not pending:
            return self.new_captcha(clean_number)
        result = self.lookup_provider.submit_captcha(clean_number, code, pending)
        if result.lookup_status == "Captcha required":
            return self.new_captcha(clean_number)
        self.pending_captcha.pop(clean_number, None)
        if result.lookup_status == "Found":
            self.save_result(result)
            return {"kind": "found", "result": asdict(result)}
        return {
            "kind": "manual",
            "number": clean_number,
            "formUrl": self.porting_lookup_url(clean_number),
            "result": asdict(result),
            "message": "Captcha was accepted, but no confirmed provider came back. Use the Porting form link and save the provider manually.",
        }

    def save_manual_result(self, clean_number: str, provider: str, raw_result: str = "") -> dict[str, Any]:
        provider = provider.strip().upper()
        raw = raw_result.strip() or f"Manual form result saved from public number-inquiry page: {provider}"
        result = LookupResult(
            clean_number,
            "Found",
            provider,
            "Yes" if provider_is_telkom(provider) else "No",
            raw,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )
        self.save_result(result)
        return {"kind": "found", "result": asdict(result)}

    def api_result_from_lookup(self, original_number: Any) -> dict[str, Any]:
        clean_number = normalize_number(original_number)
        checked_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        base = {
            "input_number": str(original_number or ""),
            "clean_number": clean_number,
            "number_type": classify_number(clean_number),
            "lookup_status": "",
            "current_provider": "",
            "telkom_service": None,
            "raw_result": "",
            "checked_at": checked_at,
            "porting_lookup_url": self.porting_lookup_url(clean_number),
        }
        if not is_supported_number(clean_number):
            return {
                **base,
                "lookup_status": "unsupported_number",
                "raw_result": "Number is not a supported South African 011, 012, or mobile number.",
            }
        if clean_number in self.cache:
            result = self.cache[clean_number]
            return self.api_result_from_lookup_result(str(original_number or ""), result)
        result_data = self.check_number(clean_number)
        if result_data.get("kind") == "found":
            return self.api_result_from_lookup_result(str(original_number or ""), LookupResult(**result_data["result"]))
        result = LookupResult(**result_data["result"])
        status = "needs_human_verification" if result_data.get("kind") == "captcha" else "needs_manual_review"
        return {
            **base,
            "lookup_status": status,
            "raw_result": result_data.get("message") or result.raw_result,
        }

    def api_result_from_lookup_result(self, original_number: str, result: LookupResult) -> dict[str, Any]:
        return {
            "input_number": original_number,
            "clean_number": result.clean_number,
            "number_type": classify_number(result.clean_number),
            "lookup_status": result.lookup_status,
            "current_provider": result.current_provider,
            "telkom_service": result.telkom == "Yes",
            "raw_result": result.raw_result,
            "checked_at": result.checked_at,
            "porting_lookup_url": self.porting_lookup_url(result.clean_number),
        }

    def create_api_job(self, numbers: list[Any], source: str = "api") -> dict[str, Any]:
        job_id = uuid.uuid4().hex
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        job = {
            "job_id": job_id,
            "source": source,
            "status": "queued",
            "total": len(numbers),
            "processed": 0,
            "telkom": 0,
            "non_telkom": 0,
            "needs_review": 0,
            "created_at": now,
            "updated_at": now,
            "message": "Queued.",
            "inputs": [str(number or "") for number in numbers],
            "results": [],
        }
        with self.api_job_lock:
            self.api_jobs[job_id] = job
            self.save_api_jobs()
        thread = threading.Thread(target=self.process_api_job, args=(job_id,), daemon=True)
        thread.start()
        return self.public_api_job(job)

    def process_api_job(self, job_id: str, delay_seconds: float = 1.0) -> None:
        with self.api_job_lock:
            job = self.api_jobs[job_id]
            job["status"] = "running"
            job["message"] = "Processing."
            job["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.save_api_jobs()
        for number in list(job["inputs"]):
            result = self.api_result_from_lookup(number)
            with self.api_job_lock:
                job = self.api_jobs[job_id]
                job["results"].append(result)
                job["processed"] = len(job["results"])
                if result["telkom_service"] is True:
                    job["telkom"] += 1
                elif result["telkom_service"] is False:
                    job["non_telkom"] += 1
                else:
                    job["needs_review"] += 1
                job["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if result["lookup_status"] == "needs_human_verification":
                    job["status"] = "needs_human_verification"
                    job["message"] = f"Paused at {result['clean_number']}: human verification is required."
                    self.save_api_jobs()
                    return
                job["message"] = f"Processed {job['processed']} of {job['total']}."
                self.save_api_jobs()
            if delay_seconds > 0:
                time.sleep(delay_seconds)
        with self.api_job_lock:
            job = self.api_jobs[job_id]
            job["status"] = "completed"
            job["message"] = "Completed."
            job["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.save_api_jobs()

    @staticmethod
    def public_api_job(job: dict[str, Any], include_results: bool = False) -> dict[str, Any]:
        public = {key: value for key, value in job.items() if key not in {"inputs", "results"}}
        if include_results:
            public["results"] = job.get("results", [])
        return public

    def get_api_job(self, job_id: str, include_results: bool = False) -> dict[str, Any] | None:
        with self.api_job_lock:
            job = self.api_jobs.get(job_id)
            if not job:
                return None
            return self.public_api_job(job, include_results=include_results)

    def api_numbers_from_file(self, file_storage) -> list[str]:
        filename = secure_filename(file_storage.filename or f"api_upload_{uuid.uuid4().hex}.csv")
        target = UPLOAD_DIR / f"api_{uuid.uuid4().hex}_{filename}"
        file_storage.save(target)
        item = {"label": target.stem, "path": str(target)}
        numbers: list[str] = []
        for source_row in self.iter_source_rows(item):
            for phone_row in self.extract_phone_rows(item, source_row):
                numbers.append(phone_row["clean_number"])
        return numbers

    def export_reports(self) -> dict[str, Any]:
        outputs = []
        for item in self.files:
            source = Path(item["path"])
            output = REPORT_DIR / f"{source.stem}_manual_checked.xlsx"
            self.export_full_report(item, output)
            outputs.append({"label": output.name, "url": url_for("download_report", filename=output.name)})
        telkom = self.export_telkom_only()
        outputs.append({"label": telkom.name, "url": url_for("download_report", filename=telkom.name)})
        route = self.export_daily_route_plan()
        outputs.append({"label": route.name, "url": url_for("download_report", filename=route.name)})
        return {"outputs": outputs}

    def result_for_export(self, clean_number: str) -> LookupResult:
        if not clean_number:
            return LookupResult(
                "",
                "Skipped",
                "",
                "Unknown",
                "No supported +27 11, +27 12, or South African mobile number was found in this row.",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )
        result = self.cache.get(clean_number)
        if result:
            return result
        return LookupResult(
            clean_number,
            "Pending",
            "",
            "Unknown",
            "Not checked yet. Resume checking in the cloud app, then export again.",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

    def export_full_report(self, item: dict[str, str], output: Path) -> None:
        source_rows = self.iter_source_rows(item)
        headers = source_rows[0]["headers"] if source_rows else []
        result_headers = [
            "Clean Number",
            "Number Type",
            "Lookup Status",
            "Current Provider",
            "Telkom Service?",
            "Raw Result",
            "Checked At",
            "Porting Lookup Link",
            "Lead Priority Score",
            "Lead Priority",
            "Visit Area",
            "Suggested Visit Order",
            "Assigned Rep",
            "Visit Status",
            "Visit Notes",
            "Next Action",
            "Score Reason",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Checked Report"
        for col, header in enumerate(list(headers) + result_headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        out_row = 2
        order_map = self.suggested_order_map([row for row in self.rows if row["source"] == item["path"]])
        for source_row in source_rows:
            extracted = self.extract_phone_rows(item, source_row)
            if not extracted:
                self.write_report_row(ws, out_row, source_row["values"], len(headers), "", "", self.result_for_export(""), {}, "")
                out_row += 1
                continue
            for phone_row in extracted:
                result = self.result_for_export(phone_row["clean_number"])
                self.write_report_row(
                    ws,
                    out_row,
                    source_row["values"],
                    len(headers),
                    phone_row["clean_number"],
                    phone_row["number_type"],
                    result,
                    phone_row,
                    order_map.get((phone_row["source"], phone_row["row"], phone_row["clean_number"]), ""),
                )
                out_row += 1

        self.finish_workbook(ws)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)

    def write_report_row(
        self,
        ws,
        out_row: int,
        values: list[Any],
        header_count: int,
        clean_number: str,
        number_type: str,
        result: LookupResult,
        phone_row: dict[str, Any],
        suggested_order: Any,
    ) -> None:
        for col, value in enumerate(values, start=1):
            ws.cell(out_row, col, value)
        first_result_col = header_count + 1
        export_values = [
            clean_number or result.clean_number,
            number_type or classify_number(result.clean_number),
            result.lookup_status,
            result.current_provider,
            result.telkom,
            result.raw_result,
            result.checked_at,
            self.porting_lookup_url(clean_number or result.clean_number),
        ]
        plan = self.lead_plan(phone_row, result)
        export_values.extend(
            [
                plan["lead_score"],
                plan["priority"],
                plan["visit_area"],
                suggested_order,
                plan["assigned_rep"],
                plan["visit_status"],
                plan["visit_notes"],
                plan["next_action"],
                plan["score_reason"],
            ]
        )
        for offset, value in enumerate(export_values):
            cell = ws.cell(out_row, first_result_col + offset, value)
            if offset == 7 and clean_number:
                cell.hyperlink = value
                cell.style = "Hyperlink"
        self.style_export_row(ws, out_row, first_result_col, result.telkom, len(export_values))

    @staticmethod
    def style_export_row(ws, row: int, first_result_col: int, telkom_value: str, width: int) -> None:
        green = PatternFill("solid", fgColor="C6EFCE")
        red = PatternFill("solid", fgColor="F4CCCC")
        amber = PatternFill("solid", fgColor="FFE699")
        fill = amber
        if telkom_value == "Yes":
            fill = green
        elif telkom_value == "No":
            fill = red
        for col in range(first_result_col, first_result_col + width):
            ws.cell(row, col).fill = fill
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)

    @staticmethod
    def finish_workbook(ws) -> None:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row, col).value or "")) for row in range(1, min(ws.max_row, 120) + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 58)

    def suggested_order_map(self, rows: list[dict[str, Any]]) -> dict[tuple[str, int, str], int]:
        unique_rows: dict[tuple[str, int, str], dict[str, Any]] = {}
        for row in rows:
            key = (row["source"], row["row"], row["clean_number"])
            unique_rows.setdefault(key, row)
        ranked = sorted(
            unique_rows.items(),
            key=lambda item: (
                -self.lead_plan(item[1], self.cache.get(item[1]["clean_number"]))["lead_score"],
                self.lead_plan(item[1], self.cache.get(item[1]["clean_number"]))["visit_area"],
                item[1].get("company", ""),
            ),
        )
        return {key: index for index, (key, _row) in enumerate(ranked, start=1)}

    def planning_rows(self) -> list[dict[str, Any]]:
        seen: set[tuple[str, int, str]] = set()
        rows: list[dict[str, Any]] = []
        order_map = self.suggested_order_map(self.rows)
        for row in self.rows:
            key = (row["source"], row["row"], row["clean_number"])
            if key in seen:
                continue
            seen.add(key)
            result = self.result_for_export(row["clean_number"])
            plan = self.lead_plan(row, result)
            rows.append(
                {
                    **row,
                    "provider": result.current_provider,
                    "telkom": result.telkom,
                    "lookup_status": result.lookup_status,
                    "raw_result": result.raw_result,
                    "checked_at": result.checked_at,
                    "suggested_order": order_map.get(key, ""),
                    **plan,
                }
            )
        return sorted(rows, key=lambda row: (row["suggested_order"] or 999999))

    def export_telkom_only(self) -> Path:
        output = REPORT_DIR / "telkom_companies_only.xlsx"
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Telkom Companies"
        headers = [
            "Source File",
            "Company",
            "Address",
            "Original Number",
            "Clean Number",
            "Number Type",
            "Current Provider",
            "Lead Priority Score",
            "Lead Priority",
            "Visit Area",
            "Suggested Visit Order",
            "Assigned Rep",
            "Visit Status",
            "Visit Notes",
            "Next Action",
            "Raw Result",
            "Checked At",
            "Porting Lookup Link",
        ]
        for col, value in enumerate(headers, start=1):
            cell = out_ws.cell(1, col, value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        out_row = 2
        seen_rows: set[tuple[str, int, str]] = set()
        order_map = self.suggested_order_map(self.rows)
        for item in self.files:
            for phone_row in [row for row in self.rows if row["source"] == item["path"]]:
                key = (phone_row["source"], phone_row["row"], phone_row["clean_number"])
                if key in seen_rows:
                    continue
                seen_rows.add(key)
                result = self.cache.get(phone_row["clean_number"])
                if not result or result.telkom != "Yes":
                    continue
                plan = self.lead_plan(phone_row, result)
                values = [
                    item["label"],
                    phone_row.get("company", ""),
                    phone_row.get("address", ""),
                    phone_row.get("original_number", ""),
                    result.clean_number,
                    phone_row.get("number_type", classify_number(result.clean_number)),
                    result.current_provider,
                    plan["lead_score"],
                    plan["priority"],
                    plan["visit_area"],
                    order_map.get(key, ""),
                    plan["assigned_rep"],
                    plan["visit_status"],
                    plan["visit_notes"],
                    plan["next_action"],
                    result.raw_result,
                    result.checked_at,
                    self.porting_lookup_url(result.clean_number),
                ]
                for col, value in enumerate(values, start=1):
                    cell = out_ws.cell(out_row, col, value)
                    if col == len(values):
                        cell.hyperlink = value
                        cell.style = "Hyperlink"
                out_row += 1

        for row in range(2, out_ws.max_row + 1):
            for col in range(1, out_ws.max_column + 1):
                out_ws.cell(row, col).fill = PatternFill("solid", fgColor="C6EFCE")
                out_ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
        self.finish_workbook(out_ws)
        out_wb.save(output)
        return output

    def export_daily_route_plan(self) -> Path:
        output = REPORT_DIR / "daily_visit_route_plan.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Route Plan"
        headers = [
            "Suggested Visit Order",
            "Lead Priority",
            "Lead Priority Score",
            "Visit Area",
            "Company",
            "Address",
            "Original Number",
            "Clean Number",
            "Number Type",
            "Current Provider",
            "Telkom Service?",
            "Lookup Status",
            "Assigned Rep",
            "Visit Status",
            "Visit Notes",
            "Next Action",
            "Score Reason",
            "Porting Lookup Link",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for out_row, row in enumerate(self.planning_rows(), start=2):
            values = [
                row["suggested_order"],
                row["priority"],
                row["lead_score"],
                row["visit_area"],
                row.get("company", ""),
                row.get("address", ""),
                row.get("original_number", ""),
                row.get("clean_number", ""),
                row.get("number_type", ""),
                row.get("provider", ""),
                row.get("telkom", ""),
                row.get("lookup_status", ""),
                row["assigned_rep"],
                row["visit_status"],
                row["visit_notes"],
                row["next_action"],
                row["score_reason"],
                self.porting_lookup_url(row.get("clean_number", "")),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(out_row, col, value)
                if col == len(values):
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
            plan_fill = PatternFill("solid", fgColor="C6EFCE" if row["priority"] == "Hot" else "FFF2CC" if row["priority"] == "Warm" else "E7E6E6")
            for col in range(1, len(values) + 1):
                ws.cell(out_row, col).fill = plan_fill
                ws.cell(out_row, col).alignment = Alignment(vertical="top", wrap_text=True)

        self.finish_workbook(ws)
        wb.save(output)
        return output


STATE = CloudState()


def authorized() -> bool:
    if not APP_PASSWORD and not API_KEYS:
        return True
    supplied = request.args.get("key") or request.headers.get("X-App-Password") or ""
    api_key = request.headers.get("X-API-Key") or ""
    auth_header = request.headers.get("Authorization") or ""
    if auth_header.lower().startswith("bearer "):
        api_key = auth_header.split(" ", 1)[1].strip()
    if api_key and any(secrets.compare_digest(api_key, key) for key in API_KEYS):
        return True
    return secrets.compare_digest(supplied, APP_PASSWORD)


@app.before_request
def require_password():
    if request.path == "/health":
        return None
    if request.path.startswith("/static"):
        return None
    if not authorized():
        return jsonify({"error": "Unauthorized. Add ?key=your-password to the URL."}), 401
    return None


PAGE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Telkom Cloud Checker</title>
  <style>
    body { margin: 0; font-family: Arial, sans-serif; background: #f6f7f9; color: #17202a; }
    main { max-width: 1120px; margin: 0 auto; padding: 28px; }
    header { display: flex; justify-content: space-between; gap: 16px; align-items: flex-start; }
    h1 { margin: 0; font-size: 28px; }
    .grid { display: grid; grid-template-columns: repeat(6, 1fr); gap: 10px; margin: 18px 0; }
    .metric, .panel { background: #fff; border: 1px solid #d9dee7; border-radius: 8px; padding: 14px; }
    .metric span { display: block; font-size: 12px; color: #657080; }
    .metric strong { display: block; font-size: 24px; margin-top: 4px; }
    .panel { margin-top: 14px; }
    .row-title { display: grid; grid-template-columns: 1.2fr .8fr .8fr; gap: 12px; }
    .label { font-size: 12px; color: #657080; margin-bottom: 4px; }
    .value { font-size: 18px; font-weight: 700; word-break: break-word; }
    button { border: 0; border-radius: 6px; background: #174ea6; color: white; padding: 11px 14px; font-weight: 700; cursor: pointer; }
    button.secondary { background: #4b5563; }
    button.danger { background: #9f1239; }
    button:disabled { background: #9aa4b2; cursor: wait; }
    input { padding: 10px; border: 1px solid #c9d1dc; border-radius: 6px; font-size: 16px; }
    .actions { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 14px; align-items: center; }
    .result { padding: 12px; border-radius: 6px; margin-top: 14px; background: #eef3ff; white-space: pre-wrap; }
    .yes { background: #d9ead3; } .no { background: #f4cccc; } .unknown { background: #ffe699; }
    .manual, .captcha { display: none; margin-top: 14px; padding: 14px; border: 1px solid #d7b945; background: #fff7d6; border-radius: 8px; }
    .captcha img { display: block; max-width: 260px; border: 1px solid #d0d0d0; background: white; margin-bottom: 10px; }
    table { width: 100%; border-collapse: collapse; font-size: 14px; }
    th, td { text-align: left; border-bottom: 1px solid #e5e7eb; padding: 8px; vertical-align: top; }
    th { background: #eef2f7; }
    a.download { display: inline-block; margin: 6px 8px 0 0; color: #174ea6; font-weight: 700; }
    @media (max-width: 850px) { .grid { grid-template-columns: 1fr 1fr; } .row-title, header { display: block; } }
  </style>
</head>
<body>
<main>
  <header>
    <div>
      <h1>Telkom Cloud Checker</h1>
      <p>Upload Excel or CSV files, auto-check 011, 012 and mobile numbers, then export reports.</p>
    </div>
    <button class="secondary" id="exportBtn">Export Reports</button>
  </header>

  <section class="panel">
    <form action="/upload{% if key %}?key={{ key|urlencode }}{% endif %}" method="post" enctype="multipart/form-data">
      <div class="label">Upload .xlsx or .csv files</div>
      <input type="file" name="files" multiple accept=".xlsx,.csv">
      <label style="display:inline-flex;gap:6px;align-items:center;margin:0 10px;"><input type="checkbox" name="auto_start" value="1" checked> Auto-check after upload</label>
      <button type="submit">Upload</button>
      <button class="danger" type="button" id="resetBtn">Reset Uploaded Files</button>
    </form>
  </section>

  <section class="grid" id="summary"></section>
  <section class="panel">
    <h2>Field Planning</h2>
    <div id="planningSummary" class="result"></div>
  </section>

  <section class="panel">
    <div class="row-title">
      <div><div class="label">Company</div><div class="value" id="company">Loading...</div></div>
      <div><div class="label">Number</div><div class="value" id="number"></div><div class="label" id="numberType"></div></div>
      <div><div class="label">File</div><div class="value" id="area"></div></div>
    </div>
    <div class="actions">
      <button id="autoBtn">Auto Check Remaining</button>
      <button class="secondary" id="stopAutoBtn">Stop Auto Check</button>
      <button id="checkBtn">Check This Number</button>
      <button class="secondary" id="formBtn">Open Porting Form</button>
      <button class="secondary" id="skipBtn">Skip For Now</button>
    </div>
    <div id="jobStatus" class="result"></div>
    <div class="captcha" id="captchaBox">
      <div class="label">Captcha from Porting site</div>
      <img id="captchaImg" alt="Captcha">
      <div class="actions">
        <input id="captchaInput" autocomplete="off" placeholder="Type captcha code">
        <button id="captchaBtn">Submit Captcha</button>
      </div>
    </div>
    <div class="manual" id="manualBox">
      <div class="label">Manual form result</div>
      <div class="actions">
        <input id="manualProvider" autocomplete="off" placeholder="Provider, e.g. TELKOM or BACKSPACE">
        <input id="manualRaw" autocomplete="off" placeholder="Optional visible message">
        <button id="manualBtn">Save Provider</button>
      </div>
    </div>
    <div id="result" class="result"></div>
    <div id="downloads"></div>
  </section>

  <section class="panel">
    <h2>Recent Checked Rows</h2>
    <table>
      <thead><tr><th>File</th><th>Company</th><th>Number</th><th>Type</th><th>Provider</th><th>Telkom Service?</th><th>Priority</th><th>Visit Area</th><th>Next Action</th></tr></thead>
      <tbody id="recent"></tbody>
    </table>
  </section>
</main>
<script>
const key = new URLSearchParams(location.search).get('key') || '';
const suffix = key ? '?key=' + encodeURIComponent(key) : '';
let current = null;
let busy = false;
let captchaNumber = '';
const formBase = 'https://www.porting.co.za/PublicWebsiteApp/#/number-inquiry?sid=smppipd4x1';
async function api(path, body) {
  const response = await fetch(path + suffix, body ? { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify(body) } : {});
  return response.json();
}
function setBusy(value) { busy = value; document.querySelectorAll('button').forEach(btn => { if (btn.id !== 'stopAutoBtn') btn.disabled = value; }); }
function metric(label, value) { return `<div class="metric"><span>${label}</span><strong>${value}</strong></div>`; }
function esc(value) {
  return String(value ?? '').replace(/[&<>"']/g, char => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[char]));
}
function withKey(url) {
  if (!key) return url;
  return url + (url.includes('?') ? '&' : '?') + 'key=' + encodeURIComponent(key);
}
function renderSummary(summary) {
  document.getElementById('summary').innerHTML = [
    metric('Files', summary.files), metric('Queue rows', summary.total_rows), metric('Unique numbers', summary.unique_numbers),
    metric('Checked', summary.checked_rows), metric('Remaining', summary.remaining_rows), metric('Telkom', summary.telkom_rows),
    metric('011 rows', summary.joburg_rows), metric('012 rows', summary.tshwane_rows), metric('Mobile rows', summary.mobile_rows),
    metric('Non-Telkom', summary.non_telkom_rows), metric('Hot leads', summary.hot_leads), metric('Warm leads', summary.warm_leads),
    metric('Cold leads', summary.cold_leads)
  ].join('');
  const areas = (summary.top_visit_areas || []).map(item => `${item.area}: ${item.count}`).join('\\n') || 'No visit areas yet.';
  document.getElementById('planningSummary').textContent = `Top visit areas\\n${areas}\\n\\nExport reports to get daily_visit_route_plan.xlsx sorted by suggested visit order.`;
  renderJob(summary.auto_job || {});
}
function renderJob(job) {
  const box = document.getElementById('jobStatus');
  const running = job.running ? 'Running' : 'Idle';
  box.className = 'result ' + (job.blocked ? 'unknown' : '');
  box.textContent = `Automatic checker: ${running}\\nChecked this run: ${job.checked_now || 0}\\nLast number: ${job.last_number || '-'}\\n${job.last_message || ''}`;
}
function renderCurrent(row, keepCaptcha = false) {
  current = row; document.getElementById('manualBox').style.display = 'none';
  if (!keepCaptcha) {
    document.getElementById('captchaBox').style.display = 'none';
    captchaNumber = '';
    document.getElementById('captchaInput').value = '';
  }
  document.getElementById('manualProvider').value = ''; document.getElementById('manualRaw').value = '';
  if (!row) {
    document.getElementById('company').textContent = 'No unchecked row available';
    document.getElementById('number').textContent = ''; document.getElementById('numberType').textContent = ''; document.getElementById('area').textContent = '';
    document.getElementById('result').textContent = 'Upload files or export reports if checking is complete.'; return;
  }
  document.getElementById('company').textContent = row.company;
  document.getElementById('number').textContent = row.clean_number;
  document.getElementById('numberType').textContent = row.number_type || '';
  document.getElementById('area').textContent = row.area;
}
function renderRecent(rows) {
  document.getElementById('recent').innerHTML = rows.slice().reverse().map(row => `
    <tr><td>${esc(row.area)}</td><td>${esc(row.company)}</td><td>${esc(row.clean_number)}</td><td>${esc(row.number_type)}</td><td>${esc(row.provider)}</td><td>${esc(row.telkom)}</td><td>${esc(row.priority)}</td><td>${esc(row.visit_area)}</td><td>${esc(row.next_action)}</td></tr>`).join('');
}
function showResult(result) {
  const box = document.getElementById('result');
  box.className = 'result ' + (result.telkom === 'Yes' ? 'yes' : result.telkom === 'No' ? 'no' : 'unknown');
  box.textContent = `${result.lookup_status}\\nProvider: ${result.current_provider || 'Unknown'}\\nTelkom: ${result.telkom}\\n${result.raw_result}`;
}
function showManual(data) {
  showResult(data.result);
  document.getElementById('manualBox').style.display = 'block';
  document.getElementById('result').textContent += '\\n\\n' + data.message;
}
function showCaptcha(data) {
  captchaNumber = data.number || (current ? current.clean_number : '');
  document.getElementById('captchaBox').style.display = 'block';
  document.getElementById('manualBox').style.display = 'none';
  document.getElementById('captchaImg').src = 'data:image/jpeg;base64,' + data.imageData;
  document.getElementById('result').className = 'result unknown';
  document.getElementById('result').textContent = data.message;
  document.getElementById('captchaInput').focus();
}
async function refresh() {
  const data = await api('/api/state');
  renderSummary(data.summary); renderCurrent(data.next, Boolean(data.pending_captcha)); renderRecent(data.recent);
  if (data.pending_captcha) showCaptcha(data.pending_captcha);
}
async function checkCurrent() {
  if (!current || busy) return; setBusy(true);
  try {
    const data = await api('/api/check', { number: current.clean_number });
    if (data.kind === 'captcha') {
      showCaptcha(data);
    } else if (data.kind === 'manual') {
      showManual(data);
    } else { showResult(data.result); await refresh(); }
  } finally { setBusy(false); }
}
async function startAuto() {
  if (busy) return; setBusy(true);
  try {
    await api('/api/auto-start', {});
    await refresh();
  } finally { setBusy(false); }
}
async function stopAuto() {
  if (busy) return; setBusy(true);
  try {
    await api('/api/auto-stop', {});
    await refresh();
  } finally { setBusy(false); }
}
function openForm() {
  const number = current ? current.clean_number : '';
  const url = number ? formBase + '&msisdn=' + encodeURIComponent(number) : formBase;
  window.open(url, '_blank', 'noopener');
}
async function submitCaptcha() {
  const number = captchaNumber || (current ? current.clean_number : '');
  if (!number || busy) return;
  const code = document.getElementById('captchaInput').value.trim();
  if (!code) return;
  setBusy(true);
  try {
    const data = await api('/api/captcha', { number, code });
    if (data.kind === 'captcha') {
      showCaptcha(data);
      document.getElementById('result').textContent = 'That code was not accepted. Try the new captcha.';
    } else if (data.kind === 'manual') {
      document.getElementById('captchaBox').style.display = 'none';
      showManual(data);
    } else {
      document.getElementById('captchaBox').style.display = 'none';
      showResult(data.result);
      await refresh();
    }
  } finally { setBusy(false); }
}
async function saveManual() {
  if (!current || busy) return;
  const provider = document.getElementById('manualProvider').value.trim();
  if (!provider) return;
  setBusy(true);
  try {
    const data = await api('/api/manual-result', {
      number: current.clean_number,
      provider,
      raw: document.getElementById('manualRaw').value.trim()
    });
    document.getElementById('manualBox').style.display = 'none';
    showResult(data.result); await refresh();
  } finally { setBusy(false); }
}
document.getElementById('checkBtn').addEventListener('click', checkCurrent);
document.getElementById('autoBtn').addEventListener('click', startAuto);
document.getElementById('stopAutoBtn').addEventListener('click', stopAuto);
document.getElementById('formBtn').addEventListener('click', openForm);
document.getElementById('captchaBtn').addEventListener('click', submitCaptcha);
document.getElementById('manualBtn').addEventListener('click', saveManual);
document.getElementById('captchaInput').addEventListener('keydown', e => { if (e.key === 'Enter') submitCaptcha(); });
document.getElementById('manualProvider').addEventListener('keydown', e => { if (e.key === 'Enter') saveManual(); });
document.getElementById('skipBtn').addEventListener('click', refresh);
document.getElementById('resetBtn').addEventListener('click', async () => { if (confirm('Remove uploaded files from this cloud app?')) { await api('/api/reset', {}); await refresh(); } });
document.getElementById('exportBtn').addEventListener('click', async () => {
  setBusy(true);
  try {
    const data = await api('/api/export', {});
    document.getElementById('downloads').innerHTML = data.outputs.map(item => `<a class="download" href="${withKey(item.url)}">${esc(item.label)}</a>`).join('');
    await refresh();
  } finally { setBusy(false); }
});
refresh();
setInterval(refresh, 5000);
</script>
</body>
</html>
"""


@app.get("/")
def index():
    return render_template_string(PAGE, key=request.args.get("key", ""))


@app.get("/health")
def health():
    return jsonify({"ok": True})


@app.post("/v1/lookup")
def v1_lookup():
    data = request.get_json(force=True)
    number = data.get("number") or data.get("msisdn") or ""
    return jsonify(STATE.api_result_from_lookup(number))


@app.post("/v1/batch")
def v1_batch():
    data = request.get_json(force=True)
    numbers = data.get("numbers") or []
    if not isinstance(numbers, list) or not numbers:
        return jsonify({"error": "Send JSON with a non-empty numbers array."}), 400
    if len(numbers) > 5000:
        return jsonify({"error": "Maximum 5000 numbers per batch job."}), 400
    return jsonify(STATE.create_api_job(numbers, source="api_batch")), 202


@app.post("/v1/files")
def v1_files():
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return jsonify({"error": "Upload one .xlsx or .csv file using form field 'file'."}), 400
    if Path(uploaded.filename).suffix.lower() not in {".xlsx", ".csv"}:
        return jsonify({"error": "Only .xlsx and .csv files are supported."}), 400
    numbers = STATE.api_numbers_from_file(uploaded)
    if not numbers:
        return jsonify({"error": "No supported 011, 012, or mobile numbers were found."}), 400
    return jsonify(STATE.create_api_job(numbers, source=f"file:{secure_filename(uploaded.filename)}")), 202


@app.get("/v1/jobs/<job_id>")
def v1_job(job_id: str):
    job = STATE.get_api_job(job_id)
    if not job:
        return jsonify({"error": "Job not found."}), 404
    return jsonify(job)


@app.get("/v1/jobs/<job_id>/results")
def v1_job_results(job_id: str):
    job = STATE.get_api_job(job_id, include_results=True)
    if not job:
        return jsonify({"error": "Job not found."}), 404
    return jsonify(job)


@app.post("/upload")
def upload():
    files = request.files.getlist("files")
    for uploaded in files:
        if uploaded and uploaded.filename:
            STATE.add_upload(uploaded)
    if request.form.get("auto_start") == "1":
        STATE.start_auto_check()
    return redirect("/" + (f"?key={request.args.get('key')}" if request.args.get("key") else ""))


@app.get("/api/state")
def api_state():
    return jsonify(
        {
            "summary": STATE.summary(),
            "next": STATE.next_row(),
            "recent": STATE.checked_rows(),
            "pending_captcha": STATE.pending_captcha_payload(),
        }
    )


@app.post("/api/check")
def api_check():
    data = request.get_json(force=True)
    return jsonify(STATE.check_number(str(data.get("number") or "")))


@app.post("/api/auto-start")
def api_auto_start():
    return jsonify({"job": STATE.start_auto_check()})


@app.post("/api/auto-stop")
def api_auto_stop():
    return jsonify({"job": STATE.stop_auto_check()})


@app.post("/api/captcha")
def api_captcha():
    data = request.get_json(force=True)
    return jsonify(STATE.submit_captcha(str(data.get("number") or ""), str(data.get("code") or "")))


@app.post("/api/manual-result")
def api_manual_result():
    data = request.get_json(force=True)
    clean_number = normalize_number(data.get("number"))
    if not clean_number:
        return jsonify({"error": "A valid number is required."}), 400
    provider = str(data.get("provider") or "").strip()
    if not provider:
        return jsonify({"error": "A provider name is required."}), 400
    return jsonify(
        STATE.save_manual_result(
            clean_number,
            provider,
            str(data.get("raw") or ""),
        )
    )


@app.post("/api/export")
def api_export():
    return jsonify(STATE.export_reports())


@app.post("/api/reset")
def api_reset():
    STATE.reset()
    return jsonify({"ok": True})


@app.get("/download/<path:filename>")
def download_report(filename: str):
    path = REPORT_DIR / secure_filename(filename)
    if not path.exists():
        return jsonify({"error": "Report not found. Export reports first."}), 404
    return send_file(path, as_attachment=True)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8765"))
    app.run(host="0.0.0.0", port=port)
