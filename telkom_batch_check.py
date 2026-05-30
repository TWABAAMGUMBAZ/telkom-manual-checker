from __future__ import annotations

import argparse
import json
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from http.cookiejar import CookieJar
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import HTTPCookieProcessor, Request, build_opener

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CRDB_REST_URL = "https://www.porting.co.za/PublicWebsiteApp/rest/"
USER_AGENT = "Mozilla/5.0 (compatible; TelkomBatchChecker/1.0)"
RESULT_HEADERS = [
    "Clean Number",
    "Number Type",
    "Lookup Status",
    "Current Provider",
    "Telkom Service?",
    "Raw Result",
    "Checked At",
]

MOBILE_PREFIXES = {
    "060",
    "061",
    "062",
    "063",
    "064",
    "065",
    "066",
    "067",
    "068",
    "069",
    "071",
    "072",
    "073",
    "074",
    "076",
    "078",
    "079",
    "081",
    "082",
    "083",
    "084",
}

TELKOM_PROVIDER_CODES = {"TELKOM", "TELKMOBL", "TELKOMMOBILE"}


@dataclass
class LookupResult:
    clean_number: str
    lookup_status: str
    current_provider: str
    telkom: str
    raw_result: str
    checked_at: str


def unchecked_result(clean_number: str, reason: str) -> LookupResult:
    return LookupResult(
        clean_number,
        "Not checked",
        "",
        "Unknown",
        reason,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


def captcha_required_result(clean_number: str) -> LookupResult:
    return LookupResult(
        clean_number,
        "Captcha required",
        "",
        "Unknown",
        "The public lookup requested a verification/captcha code. Automation stopped for this run.",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


def normalize_number(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    # Excel may store some phone-looking values as floats.
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]

    digits = re.sub(r"\D", "", text)
    if not digits:
        return ""

    if digits.startswith("0027"):
        digits = "0" + digits[4:]
    elif digits.startswith("27") and len(digits) in (11, 12):
        digits = "0" + digits[2:]

    return digits


def classify_number(clean_number: str) -> str:
    if not re.fullmatch(r"0\d{9}", clean_number or ""):
        return ""
    prefix = clean_number[:3]
    if prefix == "011":
        return "Johannesburg 011"
    if prefix == "012":
        return "Tshwane 012"
    if prefix in MOBILE_PREFIXES:
        return "Mobile"
    return ""


def is_supported_number(clean_number: str) -> bool:
    return bool(classify_number(clean_number))


class CrdbClient:
    def __init__(self, timeout: int = 30) -> None:
        self.timeout = timeout
        self.opener = build_opener(HTTPCookieProcessor(CookieJar()))
        self.puid = ""

    def request(self, method: str, path: str, payload: Any | None = None) -> dict[str, Any]:
        data = None
        if payload is not None:
            data = json.dumps(payload).encode("utf-8")
        req = Request(
            CRDB_REST_URL + path,
            data=data,
            method=method,
            headers={
                "User-Agent": USER_AGENT,
                "Content-Type": "application/json",
                "Accept": "application/json",
                "Origin": "https://www.porting.co.za",
                "Referer": "https://www.porting.co.za/PublicWebsiteApp/",
            },
        )
        with self.opener.open(req, timeout=self.timeout) as response:
            body = response.read().decode("utf-8", errors="replace")
        return json.loads(body)

    def initialize(self) -> None:
        data = self.request("GET", "publicInquiry/initializeSession")
        self.puid = str(data.get("puid") or "")

    def submit_number(self, clean_number: str) -> dict[str, Any]:
        if not self.puid:
            self.initialize()
        payload = {
            "number": clean_number,
            "captcha": "",
            "captchaEncrypt": "",
            "puid": self.puid,
        }
        return self.request("POST", "publicInquiry/submitRequest", payload)


def provider_is_telkom(provider: str) -> bool:
    parts = [part.strip().upper() for part in provider.split("/") if part.strip()]
    if not parts:
        return False
    normalized_parts = [re.sub(r"[^A-Z0-9]", "", part) for part in parts]
    return all(part in TELKOM_PROVIDER_CODES or part.startswith("TELKOM") for part in normalized_parts)


def parse_result(clean_number: str, response_data: dict[str, Any]) -> LookupResult:
    checked_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    text = str(response_data.get("msg") or "").strip()
    error_code = str(response_data.get("errorCode") or "").strip()

    if response_data.get("captchaEnabled") is True:
        return LookupResult(
            clean_number,
            "Needs review",
            "",
            "Unknown",
            "The public lookup requested captcha verification. This row was not automated.",
            checked_at,
        )

    if error_code == "NPC4046E":
        return captcha_required_result(clean_number)

    if error_code:
        raw = text or f"Lookup failed with error code {error_code}."
        return LookupResult(clean_number, "Lookup failed", "", "Unknown", raw, checked_at)

    if not text:
        raw = json.dumps(response_data, ensure_ascii=False)
        return LookupResult(clean_number, "Needs review", "", "Unknown", raw, checked_at)

    provider_match = re.search(
        r"(?:still\s+serviced\s+by|is\s+serviced\s+by)\s+([A-Z0-9_/-]+)",
        text,
        re.IGNORECASE,
    )
    if provider_match:
        provider = provider_match.group(1).upper().rstrip(".")
        return LookupResult(
            clean_number,
            "Found",
            provider,
            "Yes" if provider_is_telkom(provider) else "No",
            text,
            checked_at,
        )

    invalid_markers = [
        "invalid",
        "not a valid",
        "no record",
        "number query result",
    ]
    status = "Needs review"
    if any(marker in text.lower() for marker in invalid_markers):
        status = "No provider found"

    return LookupResult(clean_number, status, "", "Unknown", text, checked_at)


def query_number(client: CrdbClient, clean_number: str) -> LookupResult:
    try:
        response_data = client.submit_number(clean_number)
        return parse_result(clean_number, response_data)
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raw = f"HTTP {exc.code}: {exc.reason}. {body[:500]}"
    except URLError as exc:
        raw = f"Network error: {exc.reason}"
    except TimeoutError:
        raw = "Network timeout"
    except Exception as exc:
        raw = f"Lookup error: {type(exc).__name__}: {exc}"

    return LookupResult(
        clean_number,
        "Lookup failed",
        "",
        "Unknown",
        raw,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


def load_cache(path: Path) -> dict[str, LookupResult]:
    if not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    cache: dict[str, LookupResult] = {}
    for number, result in raw.items():
        lookup = LookupResult(**result)
        # Captcha and transient failures are not number-level lookup results.
        if lookup.lookup_status != "Found" or "NPC4046E" in lookup.raw_result:
            continue
        cache[number] = lookup
    return cache


def save_cache(path: Path, cache: dict[str, LookupResult]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    serializable = {number: asdict(result) for number, result in sorted(cache.items())}
    path.write_text(json.dumps(serializable, indent=2), encoding="utf-8")


def find_number_column(ws) -> int:
    header_cells = next(ws.iter_rows(min_row=1, max_row=1), [])
    candidates = {
        "contact number",
        "phone",
        "phone number",
        "telephone",
        "number",
        "msisdn",
    }
    for idx, cell in enumerate(header_cells, start=1):
        value = str(cell.value or "").strip().lower()
        if value in candidates or ("contact" in value and "number" in value):
            return idx
    raise ValueError(f"Could not find a phone number column in sheet '{ws.title}'.")


def append_result_headers(ws) -> int:
    first_result_col = ws.max_column + 1
    for offset, header in enumerate(RESULT_HEADERS):
        cell = ws.cell(row=1, column=first_result_col + offset, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    return first_result_col


def style_result_row(ws, row: int, first_result_col: int, telkom_value: str) -> None:
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="F4CCCC")
    amber = PatternFill("solid", fgColor="FFE699")

    fill = amber
    if telkom_value == "Yes":
        fill = green
    elif telkom_value == "No":
        fill = red

    for col in range(first_result_col, first_result_col + len(RESULT_HEADERS)):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)


def autosize_columns(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_length = 0
        for row in range(1, min(ws.max_row, 100) + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        width = min(max(max_length + 2, 10), 48)
        if ws.cell(row=1, column=col).value == "Raw Result":
            width = 72
        ws.column_dimensions[letter].width = width


def process_workbook(
    input_path: Path,
    output_path: Path,
    cache: dict[str, LookupResult],
    cache_path: Path,
    delay_seconds: float,
    limit: int | None = None,
    cache_only: bool = False,
) -> dict[str, int]:
    wb = load_workbook(input_path)
    ws = wb.active
    number_col = find_number_column(ws)
    first_result_col = append_result_headers(ws)
    client = CrdbClient()
    captcha_blocked = False

    stats = {
        "rows": max(ws.max_row - 1, 0),
        "checked_now": 0,
        "from_cache": 0,
        "telkom": 0,
        "non_telkom": 0,
        "review": 0,
    }

    processed = 0
    for row in range(2, ws.max_row + 1):
        original_value = ws.cell(row=row, column=number_col).value
        clean_number = normalize_number(original_value)

        if not clean_number:
            result = LookupResult(
                "",
                "Invalid input",
                "",
                "Unknown",
                f"No usable number found in source value: {original_value!r}",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )
        elif clean_number in cache:
            result = cache[clean_number]
            stats["from_cache"] += 1
        elif cache_only:
            result = unchecked_result(clean_number, "Not checked in cache-only report mode.")
        elif captcha_blocked:
            result = unchecked_result(clean_number, "Not checked because the public lookup requested captcha verification earlier in this run.")
        else:
            result = query_number(client, clean_number)
            if result.lookup_status == "Captcha required":
                captcha_blocked = True
            else:
                if result.lookup_status == "Found":
                    cache[clean_number] = result
                    save_cache(cache_path, cache)
            stats["checked_now"] += 1
            if delay_seconds > 0 and not captcha_blocked:
                time.sleep(delay_seconds)

        values = [
            result.clean_number,
            classify_number(result.clean_number),
            result.lookup_status,
            result.current_provider,
            result.telkom,
            result.raw_result,
            result.checked_at,
        ]
        for offset, value in enumerate(values):
            ws.cell(row=row, column=first_result_col + offset, value=value)
        style_result_row(ws, row, first_result_col, result.telkom)

        if result.telkom == "Yes":
            stats["telkom"] += 1
        elif result.telkom == "No":
            stats["non_telkom"] += 1
        else:
            stats["review"] += 1

        processed += 1
        if processed % 25 == 0:
            print(
                f"{input_path.name}: processed {processed}/{stats['rows']} rows "
                f"(Telkom {stats['telkom']}, non-Telkom {stats['non_telkom']}, review {stats['review']})",
                flush=True,
            )
        if limit is not None and processed >= limit:
            break

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Batch-check South African numbers against the CRDB lookup.")
    parser.add_argument("--input", action="append", required=True, help="Input .xlsx path. Repeat for multiple files.")
    parser.add_argument("--output-dir", required=True, help="Directory for checked .xlsx reports.")
    parser.add_argument("--cache", required=True, help="JSON cache path for lookup results.")
    parser.add_argument("--delay", type=float, default=1.5, help="Delay between uncached CRDB requests.")
    parser.add_argument("--limit", type=int, default=None, help="Optional row limit per workbook for testing.")
    parser.add_argument("--cache-only", action="store_true", help="Create reports from cached successful lookups without making new CRDB requests.")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    cache_path = Path(args.cache)
    cache = load_cache(cache_path)

    all_stats: dict[str, dict[str, int]] = {}
    for input_name in args.input:
        input_path = Path(input_name)
        output_path = output_dir / f"{input_path.stem}_telkom_checked.xlsx"
        stats = process_workbook(
            input_path,
            output_path,
            cache,
            cache_path,
            args.delay,
            args.limit,
            args.cache_only,
        )
        all_stats[str(output_path)] = stats
        print(f"Saved {output_path}")
        print(json.dumps(stats, indent=2))

    save_cache(cache_path, cache)
    print("SUMMARY")
    print(json.dumps(all_stats, indent=2))


if __name__ == "__main__":
    main()
